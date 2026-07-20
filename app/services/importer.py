from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.db.enums import Platform
from app.utils.links import normalize_source_link
from app.utils.regions import federal_district_for, normalize_region


@dataclass(slots=True)
class SourceCandidate:
    platform: Platform
    input_link: str
    normalized_link: str
    region: str
    federal_district: str
    category: str
    subcategory: str
    title: str = ""
    external_id: str = ""
    row_number: int = 0


@dataclass(slots=True)
class ImportErrorRow:
    row_number: int
    value: str
    reason: str


@dataclass(slots=True)
class ImportPreview:
    candidates: list[SourceCandidate]
    errors: list[ImportErrorRow]
    input_rows: int


HEADER_ALIASES = {
    "region": {"регион", "субъект", "область", "region"},
    "district": {"федеральный округ", "фо", "округ", "federal district"},
    "category": {"категория", "category"},
    "subcategory": {"подкатегория", "subcategory"},
    "link": {"ссылка", "url", "link", "ссылка на аккаунт", "ссылка на группу"},
    "tg_link": {"ссылка на тг-канал ро", "telegram", "ссылка telegram", "ссылка тг", "tg link"},
    "vk_link": {"ссылка на группу вк ро", "вконтакте", "ссылка вк", "vk link"},
    "title": {"название", "name", "title"},
    "tg_title": {"название tg", "название тг", "telegram name"},
    "vk_title": {"название вк", "vk name"},
    "tg_id": {"tg id (для парсера)", "tg id", "telegram id"},
    "vk_id": {"вк id (для парсера)", "vk id"},
    "tg_status": {"статус tg", "статус тг"},
    "vk_status": {"статус вк"},
}


def _norm_header(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _column_map(headers: list[Any]) -> dict[str, int]:
    normalized = [_norm_header(v) for v in headers]
    result: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                result[key] = idx
                break
    return result


def _value(row: list[Any], mapping: dict[str, int], key: str) -> str:
    idx = mapping.get(key)
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _status_allows(value: str) -> bool:
    status = value.strip().lower()
    return not status or status in {"рабочий", "рабочая", "работает", "active", "ok", "да"}


def _parse_rows(rows: Iterable[list[Any]]) -> ImportPreview:
    rows = iter(rows)
    try:
        headers = next(rows)
    except StopIteration:
        return ImportPreview([], [], 0)
    mapping = _column_map(headers)
    candidates: list[SourceCandidate] = []
    errors: list[ImportErrorRow] = []
    seen: set[tuple[Platform, str]] = set()
    input_rows = 0

    for row_number, row in enumerate(rows, start=2):
        input_rows += 1
        raw_region = _value(row, mapping, "region")
        region = normalize_region(raw_region)
        explicit_subcategory = _value(row, mapping, "subcategory")
        subcategory = explicit_subcategory or region
        category = (
            _value(row, mapping, "category")
            or _value(row, mapping, "district")
            or federal_district_for(region)
        )
        district = category if region else _value(row, mapping, "district")

        entries: list[tuple[str, str, str, str]] = []
        generic = _value(row, mapping, "link")
        if generic:
            entries.append((generic, _value(row, mapping, "title"), "", ""))
        tg = _value(row, mapping, "tg_link")
        if tg and _status_allows(_value(row, mapping, "tg_status")):
            entries.append((tg, _value(row, mapping, "tg_title"), _value(row, mapping, "tg_id"), "telegram"))
        vk = _value(row, mapping, "vk_link")
        if vk and _status_allows(_value(row, mapping, "vk_status")):
            entries.append((vk, _value(row, mapping, "vk_title"), _value(row, mapping, "vk_id"), "vk"))

        if not entries and any(str(v or "").strip() for v in row):
            # TXT-like single-column files parsed through CSV land here.
            first = next((str(v).strip() for v in row if v is not None and str(v).strip()), "")
            if first and row_number == 2 and not mapping:
                # Header was actually the first link. It is handled by parse_text instead.
                pass
            elif first and not mapping:
                entries.append((first, "", "", ""))

        for raw_link, title, external_id, platform_hint in entries:
            try:
                normalized = normalize_source_link(raw_link)
                if platform_hint and normalized.platform.value != platform_hint:
                    raise ValueError("Платформа ссылки не совпадает со столбцом")
                key = (normalized.platform, normalized.normalized_link)
                if key in seen:
                    continue
                seen.add(key)
                candidates.append(
                    SourceCandidate(
                        platform=normalized.platform,
                        input_link=raw_link,
                        normalized_link=normalized.normalized_link,
                        region=region,
                        federal_district=district,
                        category=category,
                        subcategory=subcategory,
                        title=title,
                        external_id=external_id,
                        row_number=row_number,
                    )
                )
            except Exception as exc:
                errors.append(ImportErrorRow(row_number, raw_link, str(exc)))

    return ImportPreview(candidates, errors, input_rows)


def parse_xlsx(path: Path) -> ImportPreview:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    try:
        if sheet is None:
            raise ValueError("XLSX workbook has no active worksheet")
        return _parse_rows(list(row) for row in sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def parse_delimited(path: Path) -> ImportPreview:
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:10000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.reader(io.StringIO(text), dialect)
    return _parse_rows(list(row) for row in reader)


def parse_text(path: Path) -> ImportPreview:
    candidates: list[SourceCandidate] = []
    errors: list[ImportErrorRow] = []
    seen: set[tuple[Platform, str]] = set()
    lines = path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
    for row_number, line in enumerate(lines, start=1):
        value = line.strip()
        if not value or value.startswith("#"):
            continue
        try:
            normalized = normalize_source_link(value)
            key = (normalized.platform, normalized.normalized_link)
            if key in seen:
                continue
            seen.add(key)
            candidates.append(
                SourceCandidate(
                    platform=normalized.platform,
                    input_link=value,
                    normalized_link=normalized.normalized_link,
                    region="",
                    federal_district="",
                    category="",
                    subcategory="",
                    row_number=row_number,
                )
            )
        except Exception as exc:
            errors.append(ImportErrorRow(row_number, value, str(exc)))
    return ImportPreview(candidates, errors, len(lines))


def parse_source_file(path: Path) -> ImportPreview:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx(path)
    if suffix in {".csv", ".tsv"}:
        return parse_delimited(path)
    if suffix == ".txt":
        return parse_text(path)
    raise ValueError("Поддерживаются XLSX, CSV, TSV и TXT")


def errors_csv(preview: ImportPreview) -> bytes:
    stream = io.StringIO()
    writer = csv.writer(stream, delimiter=";")
    writer.writerow(["Строка", "Значение", "Ошибка"])
    for row in preview.errors:
        writer.writerow([row.row_number, row.value, row.reason])
    return stream.getvalue().encode("utf-8-sig")
